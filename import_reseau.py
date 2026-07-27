#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import des comptes réseau depuis login_mot_de_passe-Reseau.xlsx
→ POST /api/serveurs (serveurs/postes)
→ POST /api/switches (switches)
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import re
import http.client
import json
import urllib.parse

EXCEL_PATH = "F:/login_mot_de_passe-Reseau.xlsx"
API_BASE   = "localhost"
API_PORT   = 8080

# ---------- helpers ----------

def api(method, path, body=None, cookie=None):
    conn = http.client.HTTPConnection(API_BASE, API_PORT, timeout=10)
    headers = {"Content-Type": "application/json"}
    if cookie:
        headers["Cookie"] = cookie
    conn.request(method, path, body=json.dumps(body) if body else None, headers=headers)
    resp = conn.getresponse()
    data = resp.read().decode("utf-8")
    conn.close()
    set_cookie = resp.getheader("Set-Cookie")
    return resp.status, data, set_cookie

def extract_ip(name):
    m = re.search(r'IP[@;][:;]?\s*(\d+\.\d+\.\d+\.\d+)', name)
    return m.group(1) if m else ""

def strip_ip(name):
    """Enlever (IP@;10.0.0.1) ou (IP@: 10.0.0.1) du nom"""
    return re.sub(r'\s*\(IP[@;][:;]?\s*\d+\.\d+\.\d+\.\d+\)', '', name).strip()

def clean_password(pwd):
    """Enlever les notes comme (0 : zero)"""
    return re.sub(r'\s*\([^)]*\)', '', pwd).strip()

# ---------- read Excel ----------

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Reseaux"]

rows = []
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    rows.append(vals)

print(f"📄 Lu {len(rows)} lignes dans Reseaux")

# ---------- parse ----------

serveurs = []   # (nom, adresseIP, login, motDePasse)
switches = []   # (nom, adresseIP, login, motDePasse, emplacement)
current_floor = ""
srv_dummy = 0
sw_dummy = 0

i = 0
while i < len(rows):
    row = rows[i]
    label = str(row[0] or "").strip()

    # Section headers
    if label == "Partie serveurs et postes de travail":
        i += 1
        continue
    if label == "Partie switching":
        i += 1
        continue
    if label in ("2ème étage", "2ème étage"):
        current_floor = "2ème étage"
        i += 1
        continue
    if label == "1er étage":
        current_floor = "1er étage"
        i += 1
        continue
    if label in ("Rez de chaussé", "Rez de chauss�"):
        current_floor = "Rez-de-chaussée"
        i += 1
        continue
    if label in ("Annexe PN",):
        current_floor = "Annexe PN"
        i += 1
        continue
    if label in ("Annexe DAEC",):
        current_floor = "Annexe DAEC"
        i += 1
        continue

    # R14 special - domain admin, skip
    if "administrateur domaine" in label.lower() or "rnp.local" in label.lower():
        print(f"  ⏭️ Ignoré (domaine): {label}")
        i += 1
        continue

    if not label or label in ("Lien/Nom ", "Lien/Nom", "Nom", "login", "Login ", ""):
        i += 1
        continue

    login_val = str(row[1] or "").strip()
    pwd_val   = clean_password(str(row[2] or "").strip())
    obs_val   = str(row[3] or "").strip()

    ip = extract_ip(label)
    nom = strip_ip(label)

    # Determine if serveur or switch based on section
    # After "Partie switching", everything is a switch
    # Before, everything is a serveur
    if label.startswith("Switch") or label.startswith("switch"):
        # UQ_SwitchReseau_AdresseMAC rejects '' — use unique dummy
        adr_mac = f"IMPORT_DUMMY_{sw_dummy:04d}" if not row[4] else str(row[4]).strip()
        sw_dummy += 1
        switches.append((nom, ip, login_val, pwd_val, current_floor, adr_mac))
        print(f"  \U0001f500 Switch: {nom} | IP={ip} | login={login_val or '*'}")
    else:
        # UQ_Serveur_AdresseIP rejects '' — use unique dummy
        if not ip:
            srv_dummy += 1
            ip = f"0.0.0.{srv_dummy}"
        serveurs.append((nom, ip, login_val, pwd_val))
        print(f"  \U0001f5a5️  Serveur: {nom} | IP={ip} | login={login_val}")

    i += 1

print(f"\n\U0001f4ca Parsing: {len(serveurs)} serveurs, {len(switches)} switches")

# ---------- login ----------

if serveurs or switches:
    print("\n\U0001f510 Connexion \U0001f511 l'API...")
    status, data, cookie = api("POST", "/api/login", {"login": "admin", "motDePasse": "admin123"})
    if status != 200:
        print(f"❌ Échec login: {status} {data}")
        exit(1)
    session_cookie = ""
    if cookie:
        session_cookie = cookie.split(";")[0]
    else:
        print("⚠️  Set-Cookie absent, tentative sans cookie...")
        resp_json = json.loads(data)
        if resp_json.get("succes") != "true":
            print(f"❌ Login refusé: {data}")
            exit(1)

    print(f"✅ Connecté")
    if not session_cookie:
        conn = http.client.HTTPConnection(API_BASE, API_PORT, timeout=5)
        conn.request("POST", "/api/login", body=json.dumps({"login": "admin", "motDePasse": "admin123"}), headers={"Content-Type": "application/json"})
        resp = conn.getresponse()
        resp.read()
        sc = resp.getheader("Set-Cookie", "")
        if sc:
            session_cookie = sc.split(";")[0]
        conn.close()

    if not session_cookie:
        print("❌ Impossible d'obtenir le cookie de session")
        exit(1)

    print(f"\U0001f36a Cookie: {session_cookie}")

    # ---------- import serveurs ----------

    print(f"\n\U0001f5a5️  Import des {len(serveurs)} serveurs...")
    for nom, ip, login, pwd in serveurs:
        if not login:
            print(f"  ⏭️ {nom}: pas de login, ignoré")
            continue
        if not pwd:
            print(f"  ⏭️ {nom}: pas de mot de passe, ignoré")
            continue
        payload = {
            "nom": nom,
            "adresseIP": ip,
            "login": login,
            "motDePasse": pwd
        }
        status, data, _ = api("POST", "/api/serveurs", payload, session_cookie)
        if status == 201:
            print(f"  ✅ {nom} → {data}")
        elif status == 400 and "existe déjà" in data.lower():
            print(f"  ⏭️ {nom}: déjà existant")
        else:
            print(f"  ❌ {nom}: HTTP {status} {data}")

    # ---------- import switches ----------

    print(f"\n\U0001f500 Import des {len(switches)} switches...")
    for nom, ip, login, pwd, emplacement, mac in switches:
        if not pwd:
            print(f"  ⏭️ {nom}: pas de mot de passe, ignoré")
            continue
        payload = {
            "nom": nom,
            "adresseMac": mac,
            "emplacement": emplacement,
            "login": login if login else "",
            "motDePasse": pwd
        }
        status, data, _ = api("POST", "/api/switches", payload, session_cookie)
        if status == 201:
            print(f"  ✅ {nom} → {data}")
        elif status == 400 and "existe déjà" in data.lower():
            print(f"  ⏭️ {nom}: déjà existant")
        else:
            print(f"  ❌ {nom}: HTTP {status} {data}")

print("\n✅ Import terminé !")